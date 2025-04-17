# import tkinter as tk
# from tkinter import simpledialog
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from webdriver_manager.chrome import ChromeDriverManager
# from bs4 import BeautifulSoup
# import pandas as pd
# import time
# import re
# import datetime

# options = Options()
# options.add_experimental_option("detach", True)
# driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)


# wait = WebDriverWait(driver, 15)  # Wait up to 15 seconds


# all_products = []

# try:
#     driver.get("https://www.noon.com/saudi-ar/p-105882/")
    
    
#     wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.ProductBoxLinkHandler_linkWrapper__b0qZ9")))
    
#     current_page = 1
#     max_pages = 2
    
#     while current_page <= max_pages:
#         print(f"🔄 Processing page {current_page}...")
        
       
#         driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
#         time.sleep(2)  # Small pause for any lazy-loaded content
        
        
#         wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.ProductBoxLinkHandler_linkWrapper__b0qZ9")))
        
        
#         soup = BeautifulSoup(driver.page_source, 'html.parser')
#         product_cards = soup.find_all("div", "ProductBoxLinkHandler_linkWrapper__b0qZ9")
        
#         if not product_cards:
#             print("⚠️ No products found on page. Stopping.")
#             break
            
#         for card in product_cards:
#             product_name = card.find('h2', {"data-qa": "product-name"})
#             price_tag = card.find("strong", class_="Price_amount__2sXa7")
#             old_price = card.find('span', class_="Price_oldPrice__ZqD8B")
#             discount_percent = card.find('span', "PriceDiscount_discount__1ViHb PriceDiscount_pBox__eWMKb")
#             link = card.find('a', class_='ProductBoxLinkHandler_productBoxLink__FPhjp')
#             href = link.get('href') if link else None
#             image = card.find('img', class_="ProductImageCarousel_productImage__jtsOn")
#             src = image.get('src') if image else None
            
#             product = {
#                 "اسم المنتج": product_name.get('title').strip() if product_name else None,
#                 "السعر": price_tag.get_text(strip=True) if price_tag else None,
#                 "السعر قبل الخصم": old_price.get_text(strip=True) if old_price else None,
#                 "نسبة الخصم": discount_percent.get_text(strip=True) if discount_percent else None,
#                 "رابط المنتج": f"https://www.noon.com{href}" if href else None,
#                 "صورة المنتج": src if src else None
#             }
#             all_products.append(product)
        
#         print(f"Page {current_page} scraped with {len(product_cards)} products.")
        
#         # Try to navigate to next page
#         if current_page < max_pages:
#             try:
#                 next_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[aria-label="Next page"]')))
#                 driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", next_button)
#                 time.sleep(1)  # Small pause for any animations
#                 next_button.click()
                
                
#                 wait.until(lambda d: f"page={current_page + 1}" in d.current_url)
#                 current_page += 1
#             except Exception as e:
#                 print(f"❌ Could not navigate to next page: {str(e)}")
#                 break
#         else:
#             break
            
# except Exception as e:
#     print(f"❌ An error occurred: {str(e)}")
# finally:
#     driver.quit()


# if all_products:
#     df = pd.DataFrame(all_products)
    
#     # Clean and process data
#     df.drop_duplicates(inplace=True)
#     df.dropna(subset=["اسم المنتج", "السعر"], inplace=True)

#     def clean_price(value):
#         if value:
#             cleaned = re.sub(r"[^\d.]", "", value.replace(",", ""))
#             return float(cleaned) if cleaned else None
#         return None

#     df["السعر"] = df["السعر"].apply(clean_price)
#     df["السعر قبل الخصم"] = df["السعر قبل الخصم"].apply(clean_price)

#     def calculate_discount(row):
#         if row["السعر قبل الخصم"] and row["السعر"]:
#             discount_amount = row["السعر قبل الخصم"] - row["السعر"]
#             discount_percentage = (discount_amount / row["السعر قبل الخصم"]) * 100 if row["السعر قبل الخصم"] else 0
#             return pd.Series([discount_amount, discount_percentage])
#         return pd.Series([None, None])

#     df[["مقدار الخصم", "نسبة الخصم المحسوبة"]] = df.apply(calculate_discount, axis=1)
#     df.drop(columns=["مقدار الخصم"], inplace=True)
#     df.sort_values(by=["نسبة الخصم المحسوبة"], ascending=False, inplace=True)
#     df.reset_index(drop=True, inplace=True)

    
#     timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
#     filename = f"noon_products_{timestamp}.xlsx"
#     df.to_excel(filename, index=False)
#     print(f" Success! Scraped {len(df)} products. Data saved to '{filename}'.")
# else:
#     print("No products were scraped.")

# import tkinter as tk
# from tkinter import simpledialog
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from webdriver_manager.chrome import ChromeDriverManager
# from bs4 import BeautifulSoup
# import pandas as pd
# import time
# import re
# import datetime

# options = Options()
# options.add_experimental_option("detach", True)
# driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# wait = WebDriverWait(driver, 15)  # Wait up to 15 seconds

# all_products = []

# def get_product_description(url):
#     try:
#         # Open a new window
#         driver.execute_script("window.open('');")
#         # Switch to the new window
#         driver.switch_to.window(driver.window_handles[1])
#         driver.get(url)
        
#         # Scroll down to load the page content
#         for _ in range(3):  # Scroll 3 times
#             driver.execute_script("window.scrollBy(0, 500)")
#             time.sleep(1)
        
#         try:
#             # Wait for the description element to be present
#             description_element = wait.until(
#                 EC.presence_of_element_located((By.CSS_SELECTOR, "div.OverviewTab_overviewDescriptionCtr__d5ELj"))
#             )
#             description = description_element.text.strip()
#         except Exception as e:
#             print(f"Could not find description on {url}: {str(e)}")
#             description = None
        
#         # Close the current window
#         driver.close()
#         # Switch back to the original window
#         driver.switch_to.window(driver.window_handles[0])
        
#         return description
#     except Exception as e:
#         print(f"Error processing {url}: {str(e)}")
#         driver.switch_to.window(driver.window_handles[0])
#         return None

# try:
#     driver.get("https://www.noon.com/saudi-ar/p-105882/")
    
#     wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.ProductBoxLinkHandler_linkWrapper__b0qZ9")))
    
#     current_page = 1
#     max_pages = 2
    
#     while current_page <= max_pages:
#         print(f"🔄 Processing page {current_page}...")
        
#         driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
#         time.sleep(2)  # Small pause for any lazy-loaded content
        
#         wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.ProductBoxLinkHandler_linkWrapper__b0qZ9")))
        
#         soup = BeautifulSoup(driver.page_source, 'html.parser')
#         product_cards = soup.find_all("div", "ProductBoxLinkHandler_linkWrapper__b0qZ9")
        
#         if not product_cards:
#             print("⚠️ No products found on page. Stopping.")
#             break
            
#         for card in product_cards:
#             product_name = card.find('h2', {"data-qa": "product-name"})
#             price_tag = card.find("strong", class_="Price_amount__2sXa7")
#             old_price = card.find('span', class_="Price_oldPrice__ZqD8B")
#             discount_percent = card.find('span', "PriceDiscount_discount__1ViHb PriceDiscount_pBox__eWMKb")
#             link = card.find('a', class_='ProductBoxLinkHandler_productBoxLink__FPhjp')
#             href = link.get('href') if link else None
#             image = card.find('img', class_="ProductImageCarousel_productImage__jtsOn")
#             src = image.get('src') if image else None
            
#             product_url = f"https://www.noon.com{href}" if href else None
#             product_description = get_product_description(product_url) if product_url else None
            
#             product = {
#                 "اسم المنتج": product_name.get('title').strip() if product_name else None,
#                 "السعر": price_tag.get_text(strip=True) if price_tag else None,
#                 "السعر قبل الخصم": old_price.get_text(strip=True) if old_price else None,
#                 "نسبة الخصم": discount_percent.get_text(strip=True) if discount_percent else None,
#                 "رابط المنتج": product_url,
#                 "صورة المنتج": src if src else None,
#                 "وصف المنتج": product_description
#             }
#             all_products.append(product)
        
#         print(f"Page {current_page} scraped with {len(product_cards)} products.")
        
#         # Try to navigate to next page
#         if current_page < max_pages:
#             try:
#                 next_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[aria-label="Next page"]')))
#                 driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", next_button)
#                 time.sleep(1)  # Small pause for any animations
#                 next_button.click()
                
#                 wait.until(lambda d: f"page={current_page + 1}" in d.current_url)
#                 current_page += 1
#             except Exception as e:
#                 print(f"❌ Could not navigate to next page: {str(e)}")
#                 break
#         else:
#             break
            
# except Exception as e:
#     print(f"❌ An error occurred: {str(e)}")
# finally:
#     driver.quit()

# if all_products:
#     df = pd.DataFrame(all_products)
    
#     # Clean and process data
#     df.drop_duplicates(inplace=True)
#     df.dropna(subset=["اسم المنتج", "السعر"], inplace=True)

#     def clean_price(value):
#         if value:
#             cleaned = re.sub(r"[^\d.]", "", value.replace(",", ""))
#             return float(cleaned) if cleaned else None
#         return None

#     df["السعر"] = df["السعر"].apply(clean_price)
#     df["السعر قبل الخصم"] = df["السعر قبل الخصم"].apply(clean_price)

#     def calculate_discount(row):
#         if row["السعر قبل الخصم"] and row["السعر"]:
#             discount_amount = row["السعر قبل الخصم"] - row["السعر"]
#             discount_percentage = (discount_amount / row["السعر قبل الخصم"]) * 100 if row["السعر قبل الخصم"] else 0
#             return pd.Series([discount_amount, discount_percentage])
#         return pd.Series([None, None])

#     df[["مقدار الخصم", "نسبة الخصم المحسوبة"]] = df.apply(calculate_discount, axis=1)
#     df.drop(columns=["مقدار الخصم"], inplace=True)
#     df.sort_values(by=["نسبة الخصم المحسوبة"], ascending=False, inplace=True)
#     df.reset_index(drop=True, inplace=True)

#     timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
#     filename = f"noon_products_{timestamp}.xlsx"
#     df.to_excel(filename, index=False)
#     print(f" Success! Scraped {len(df)} products. Data saved to '{filename}'.")
# else:
#     print("No products were scraped.")
import tkinter as tk
from tkinter import simpledialog
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import pandas as pd
import time
import re
import datetime
import os
import requests
from PIL import Image
from io import BytesIO
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage

options = Options()
options.add_experimental_option("detach", True)
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

wait = WebDriverWait(driver, 15)  # Wait up to 15 seconds

all_products = []
downloaded_images = {}  # To store downloaded image data

def download_image(url, product_id):
    if not url or url in downloaded_images:
        return None
    
    try:
        response = requests.get(url, stream=True)
        response.raise_for_status()
        
        # Open the image to get dimensions
        img = Image.open(BytesIO(response.content))
        
        # Resize if necessary (optional)
        max_size = (200, 200)
        img.thumbnail(max_size, Image.Resampling.LANCZOS)
        
        # Store the image data
        img_byte_arr = BytesIO()
        img.save(img_byte_arr, format='PNG')
        downloaded_images[url] = img_byte_arr.getvalue()
        
        return img_byte_arr.getvalue()
    except Exception as e:
        print(f"Error downloading image {url}: {str(e)}")
        return None

def get_product_details(url):
    details = {
        "description": None,
        "overview_items": []  # Store all li elements
    }
    
    try:
        # Open a new window
        driver.execute_script("window.open('');")
        # Switch to the new window
        driver.switch_to.window(driver.window_handles[1])
        driver.get(url)
        
        # Scroll down to load the page content
        for _ in range(3):
            driver.execute_script("window.scrollBy(0, 500)")
            time.sleep(1)
        
        try:
            # Get description
            description_element = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div.OverviewTab_overviewDescriptionCtr__d5ELj"))
            )
            details["description"] = description_element.text.strip()
        except Exception as e:
            print(f"Could not find description on {url}: {str(e)}")
        
        try:
            # Get all li elements from the overview container
            overview_container = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div.OverviewTab_container__2ewCs"))
            )
            
            # Parse with BeautifulSoup
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            overview_ul = soup.select_one("div.OverviewTab_container__2ewCs ul")
            
            if overview_ul:
                li_elements = overview_ul.find_all('li')
                details["overview_items"] = [li.get_text(strip=True) for li in li_elements if li.get_text(strip=True)]
                
        except Exception as e:
            print(f"Could not find overview container on {url}: {str(e)}")
        
        # Close the current window
        driver.close()
        # Switch back to the original window
        driver.switch_to.window(driver.window_handles[0])
        
    except Exception as e:
        print(f"Error processing {url}: {str(e)}")
        if len(driver.window_handles) > 1:
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
    
    return details

try:
    driver.get("https://www.noon.com/saudi-ar/p-105882/")
    
    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.ProductBoxLinkHandler_linkWrapper__b0qZ9")))
    
    current_page = 1
    max_pages = 2
    
    while current_page <= max_pages:
        print(f"🔄 Processing page {current_page}...")
        
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)  # Small pause for any lazy-loaded content
        
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.ProductBoxLinkHandler_linkWrapper__b0qZ9")))
        
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        product_cards = soup.find_all("div", "ProductBoxLinkHandler_linkWrapper__b0qZ9")
        
        if not product_cards:
            print("⚠️ No products found on page. Stopping.")
            break
            
        for i, card in enumerate(product_cards):
            product_name = card.find('h2', {"data-qa": "product-name"})
            price_tag = card.find("strong", class_="Price_amount__2sXa7")
            old_price = card.find('span', class_="Price_oldPrice__ZqD8B")
            discount_percent = card.find('span', "PriceDiscount_discount__1ViHb PriceDiscount_pBox__eWMKb")
            link = card.find('a', class_='ProductBoxLinkHandler_productBoxLink__FPhjp')
            href = link.get('href') if link else None
            image = card.find('img', class_="ProductImageCarousel_productImage__jtsOn")
            src = image.get('src') if image else None
            
            product_url = f"https://www.noon.com{href}" if href else None
            product_details = get_product_details(product_url) if product_url else {}
            
            # Download the image
            image_data = None
            if src:
                image_data = download_image(src, f"product_{current_page}_{i}")
            
            # Create product dictionary
            product = {
                "اسم المنتج": product_name.get('title').strip() if product_name else None,
                "السعر": price_tag.get_text(strip=True) if price_tag else None,
                "السعر قبل الخصم": old_price.get_text(strip=True) if old_price else None,
                "نسبة الخصم": discount_percent.get_text(strip=True) if discount_percent else None,
                "رابط المنتج": product_url,
                "صورة المنتج URL": src if src else None,
                "وصف المنتج": product_details.get("description"),
                "تفاصيل المنتج": "\n• ".join([""] + product_details.get("overview_items", [])),  # Bullet points
                "Image Data": image_data
            }
            
            all_products.append(product)
        
        print(f"Page {current_page} scraped with {len(product_cards)} products.")
        
        # Try to navigate to next page
        if current_page < max_pages:
            try:
                next_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[aria-label="Next page"]')))
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", next_button)
                time.sleep(1)  # Small pause for any animations
                next_button.click()
                
                wait.until(lambda d: f"page={current_page + 1}" in d.current_url)
                current_page += 1
            except Exception as e:
                print(f"❌ Could not navigate to next page: {str(e)}")
                break
        else:
            break
            
except Exception as e:
    print(f"❌ An error occurred: {str(e)}")
finally:
    driver.quit()

if all_products:
    df = pd.DataFrame(all_products)
    
    # Clean and process data
    df.drop_duplicates(inplace=True)
    df.dropna(subset=["اسم المنتج", "السعر"], inplace=True)

    def clean_price(value):
        if value:
            cleaned = re.sub(r"[^\d.]", "", value.replace(",", ""))
            return float(cleaned) if cleaned else None
        return None

    df["السعر"] = df["السعر"].apply(clean_price)
    df["السعر قبل الخصم"] = df["السعر قبل الخصم"].apply(clean_price)

    def calculate_discount(row):
        if row["السعر قبل الخصم"] and row["السعر"]:
            discount_amount = row["السعر قبل الخصم"] - row["السعر"]
            discount_percentage = (discount_amount / row["السعر قبل الخصم"]) * 100 if row["السعر قبل الخصم"] else 0
            return pd.Series([discount_amount, discount_percentage])
        return pd.Series([None, None])

    df[["مقدار الخصم", "نسبة الخصم المحسوبة"]] = df.apply(calculate_discount, axis=1)
    df.drop(columns=["مقدار الخصم"], inplace=True)
    df.sort_values(by=["نسبة الخصم المحسوبة"], ascending=False, inplace=True)
    df.reset_index(drop=True, inplace=True)

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"noon_products_{timestamp}.xlsx"
    
    # First save to Excel without images
    
    df=df.drop(columns=["نسبة الخصم المحسوبة"])
    df.to_excel(filename, index=False)
    
    # Now add images to the Excel file
    
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Add a header for the image column
    image_col = len(df.columns) + 1  # Place images in the next available column
    image_col_letter = openpyxl.utils.get_column_letter(image_col)
    ws[f"{image_col_letter}1"] = "صورة المنتج"
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(df)+1), start=1):
        img_data = df.at[idx-1, "Image Data"]
        if img_data:
            try:
                img = ExcelImage(BytesIO(img_data))
                img.anchor = f'{image_col_letter}{idx+1}'  # Place image in the image column
                img.width = 100  # Set width
                img.height = 100  # Set height
                ws.add_image(img)
                # Adjust row height to fit image
                ws.row_dimensions[idx+1].height = 80
            except Exception as e:
                print(f"Error adding image for row {idx}: {str(e)}")
    
    # Adjust column width for image column
    ws.column_dimensions[image_col_letter].width = 15
    
    # Save the workbook with images
    wb.save(filename)
    
    print(f"✅ Success! Scraped {len(df)} products with details and images. Data saved to '{filename}'.")
else:
    print("No products were scraped.")